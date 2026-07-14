#!/usr/bin/env python3
"""Genera la plantilla de seguimiento de señales XAU/USD (Excel con fórmulas).

Crea `Seguimiento_XAUUSD.xlsx` con dos hojas:
  * «Operaciones»: la tabla que rellenas (una fila por señal).
  * «Resumen»: métricas que se calculan solas (acierto, Profit Factor, rachas…).

Uso:  python plantillas/generar_plantilla.py
Requiere: openpyxl.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

FILAS = 500  # capacidad de operaciones.
ORO = "E8B923"
GRIS = "141A24"
GRIS_CLARO = "F2F2F2"

fino = Side(style="thin", color="D9D9D9")
borde = Border(left=fino, right=fino, top=fino, bottom=fino)


def _cab(celda, texto, ancho=None, ws=None):
    celda.value = texto
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill("solid", fgColor=GRIS)
    celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    celda.border = borde


def construir() -> Workbook:
    wb = Workbook()

    # ---------------- Hoja Operaciones ----------------
    op = wb.active
    op.title = "Operaciones"
    encabezados = [
        ("Nº", 6), ("Fecha", 12), ("Dirección", 11), ("Entrada", 10),
        ("Stop Loss", 10), ("TP objetivo", 11), ("Confianza %", 11),
        ("Resultado", 12), ("R obtenido", 11), ("Notas", 34),
    ]
    for col, (texto, ancho) in enumerate(encabezados, start=1):
        _cab(op.cell(row=1, column=col), texto)
        op.column_dimensions[op.cell(row=1, column=col).column_letter].width = ancho

    # Columnas auxiliares (rachas) — se ocultan.
    op.cell(row=1, column=12, value="_gan").font = Font(color="BBBBBB", size=8)
    op.cell(row=1, column=13, value="_per").font = Font(color="BBBBBB", size=8)

    # Fila de ejemplo (se puede borrar).
    ejemplo = [1, "2026-07-14", "Compra", 4025.80, 4015.02, 4047.37, 67, "Ganada", 1.7,
               "Ejemplo — puedes borrar esta fila"]
    for col, val in enumerate(ejemplo, start=1):
        c = op.cell(row=2, column=col, value=val)
        c.border = borde
        c.font = Font(italic=True, color="888888")

    # Fórmulas por fila:
    #   · Nº automático (solo si la fila tiene fecha).
    #   · Columnas auxiliares L/M: racha consecutiva de ganadas / perdidas.
    #     Cada fila suma 1 a la racha anterior de su tipo, o la reinicia a 0.
    #     N(...) convierte texto/vacío en 0, así la primera fila y las cabeceras
    #     no dan error.
    for r in range(2, FILAS + 2):
        prev = r - 1
        op.cell(row=r, column=1).value = f'=IF(B{r}="","",ROW()-1)'
        op.cell(row=r, column=12).value = f'=IF(H{r}="Ganada",N(L{prev})+1,0)'
        op.cell(row=r, column=13).value = f'=IF(H{r}="Perdida",N(M{prev})+1,0)'

    op.column_dimensions["L"].hidden = True
    op.column_dimensions["M"].hidden = True
    op.freeze_panes = "A2"

    # Validaciones desplegables.
    dv_dir = DataValidation(type="list", formula1='"Compra,Venta"', allow_blank=True)
    dv_res = DataValidation(type="list", formula1='"Ganada,Perdida,BE,En curso"', allow_blank=True)
    op.add_data_validation(dv_dir)
    op.add_data_validation(dv_res)
    dv_dir.add(f"C2:C{FILAS+1}")
    dv_res.add(f"H2:H{FILAS+1}")

    # ---------------- Hoja Resumen ----------------
    rs = wb.create_sheet("Resumen", 0)
    rs.column_dimensions["A"].width = 26
    rs.column_dimensions["B"].width = 16

    t = rs.cell(row=1, column=1, value="SEGUIMIENTO DE SEÑALES XAU/USD — Cuenta demo")
    t.font = Font(bold=True, size=14, color=ORO)
    rs.merge_cells("A1:B1")
    d = rs.cell(row=2, column=1,
                value="Herramienta de análisis, no asesoramiento financiero. Rellena solo la hoja «Operaciones».")
    d.font = Font(italic=True, size=9, color="888888")
    rs.merge_cells("A2:B2")

    fil = "Operaciones"
    métricas = [
        ("Operaciones cerradas", f"=COUNT({fil}!I2:I{FILAS+1})", "0"),
        ("Ganadas", f'=COUNTIF({fil}!H2:H{FILAS+1},"Ganada")', "0"),
        ("Perdidas", f'=COUNTIF({fil}!H2:H{FILAS+1},"Perdida")', "0"),
        ("Break-even (BE)", f'=COUNTIF({fil}!H2:H{FILAS+1},"BE")', "0"),
        ("En curso (abiertas)", f'=COUNTIF({fil}!H2:H{FILAS+1},"En curso")', "0"),
        ("% de ACIERTO", f'=IFERROR(B6/(B6+B7),0)', "0.0%"),
        ("Profit Factor", f'=IFERROR(SUMIF({fil}!I2:I{FILAS+1},">0")/-SUMIF({fil}!I2:I{FILAS+1},"<0"),0)', "0.00"),
        ("Expectancy (R media)", f'=IFERROR(SUM({fil}!I2:I{FILAS+1})/COUNT({fil}!I2:I{FILAS+1}),0)', "0.000"),
        ("Resultado total (R)", f"=SUM({fil}!I2:I{FILAS+1})", "+0.0;-0.0"),
        ("Racha ganadora máx", f"=MAX({fil}!L2:L{FILAS+1})", "0"),
        ("Racha perdedora máx", f"=MAX({fil}!M2:M{FILAS+1})", "0"),
    ]
    fila = 4
    rs.cell(row=fila, column=1, value="MÉTRICA").font = Font(bold=True, color="FFFFFF")
    rs.cell(row=fila, column=1).fill = PatternFill("solid", fgColor=GRIS)
    rs.cell(row=fila, column=2, value="VALOR").font = Font(bold=True, color="FFFFFF")
    rs.cell(row=fila, column=2).fill = PatternFill("solid", fgColor=GRIS)
    rs.cell(row=fila, column=1).border = borde
    rs.cell(row=fila, column=2).border = borde
    fila += 1
    for i, (nombre, formula, fmt) in enumerate(métricas):
        a = rs.cell(row=fila, column=1, value=nombre)
        b = rs.cell(row=fila, column=2, value=formula)
        a.border = borde
        b.border = borde
        b.number_format = fmt
        b.alignment = Alignment(horizontal="center")
        relleno = GRIS_CLARO if i % 2 == 0 else "FFFFFF"
        a.fill = PatternFill("solid", fgColor=relleno)
        b.fill = PatternFill("solid", fgColor=relleno)
        if nombre == "% de ACIERTO":
            b.font = Font(bold=True, size=12, color="1B7A3D")
        fila += 1

    nota = rs.cell(row=fila + 1, column=1,
                   value="Cómo usarla: por cada señal del email, apunta una fila en «Operaciones». "
                         "Cuando cierre, marca «Ganada/Perdida/BE» y escribe el R obtenido "
                         "(+1.7 si llegó a todos los TP, +0.5 si salió en break-even, -1 si saltó el stop).")
    nota.font = Font(italic=True, size=9, color="888888")
    nota.alignment = Alignment(wrap_text=True, vertical="top")
    rs.merge_cells(start_row=fila + 1, start_column=1, end_row=fila + 4, end_column=2)

    return wb


if __name__ == "__main__":
    import os
    salida = os.path.join(os.path.dirname(__file__), "Seguimiento_XAUUSD.xlsx")
    construir().save(salida)
    print("Plantilla generada en:", salida)
